from openpyxl import load_workbook

wb = load_workbook("D:/Files/Study/code/DataProcessing/assessment_models/Data_processsing/2023_MCM_Problem_C_Data.xlsx", data_only=True)
ws = wb.active

for r in range(1, 6):
    print(r, [c.value for c in ws[r]])
